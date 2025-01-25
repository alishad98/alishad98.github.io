import win32com.client

import openpyxl

import pyperclip

import os

import time

import pandas as pd

from SAP_Login_Class import SapGui

 

# Function to send an email using Microsoft Outlook

def send_email(subject, body, to):

    outlook = win32com.client.Dispatch("Outlook.Application")  # Create an Outlook application object

    mail = outlook.CreateItem(0)  # Create a new mail item (0 represents an email)

    mail.Subject = subject  # Set the subject of the email

    mail.Body = body  # Set the body of the email

    mail.To = to  # Set the recipient(s) of the email

    mail.Send()  # Send the email

 

try:

    import time

    start_time = time.time()  # Starts Timer to count how long code takes to run

 

    # Read data from the control sheet and create a dataframe

    df_control = pd.read_excel(r"\\groups-server\groups\Design\Dept Only\Cost Analysts\PlanVsActual DATA\NowCast Data Auto\NowCast Control Workbook.xlsx")

 

    # Extracting values from the control sheet for various parameters

    Dir = df_control.iloc[26, 1]

    ProdOExpName = df_control.iloc[45, 3]

    ProdOSD = df_control.iloc[3, 1]

    ProdOEd = df_control.iloc[4, 1]

    OpExpName = df_control.iloc[12, 1]

    OpFiltDate = df_control.iloc[10, 1]

    ComExpName = df_control.iloc[14, 1]

    ComFiltDate = df_control.iloc[11, 1]

    ExOPExpName = df_control.iloc[21, 1]

    ExOpSD = df_control.iloc[3, 1]

    ExOpEd = df_control.iloc[4, 1]

    SetExpName = df_control.iloc[52, 2]

    ZMudExpName = df_control.iloc[41, 3]

    ZMudSD = df_control.iloc[57, 1]

    ZMudEd = df_control.iloc[58, 1]

    MatExpName = df_control.iloc[43, 3]

    MatMSD = df_control.iloc[55, 1]

    MatMEd = df_control.iloc[56, 1]

    MB51ExpName = df_control.iloc[39, 3]

    ZLLRExpName = df_control.iloc[47, 3]

 

    # Create an instance of the SAP_GUI class

    SAP_OBJ = SapGui()

 

    # Connection to SAP and login

    def ConnectionTest():

        try:

            SAP_OBJ.connectsap()  # Try connecting to SAP using the SAP_GUI object

        except:

            SAP_OBJ.saplogin()  # If the connection fails, perform a new login using the SAP_GUI object

 

    ConnectionTest()  # Call the ConnectionTest function to establish a connection to SAP

    # Production Order Extract Block

 

    # The following lines of code interact with SAP GUI to extract production orders data.

    print('Test1')

 

#Start of WT2 Recharges (Wind Tunnel) daily export

# Access relevant SAP transaction, type in company code = 1000

    SAP_OBJ.session.findById("wnd[0]/tbar[0]/okcd").text = "S_ALR_87013019"

    SAP_OBJ.session.findById("wnd[0]").sendVKey(0)

    SAP_OBJ.session.findById("wnd[0]/usr/txt$6-KOKRS").text = "1000"

 

    # Function to filter values starting with specific prefixes

    def filter_values(sheet, prefixes):

        matching_values = []

        for row in sheet.iter_rows(values_only=True):

            for cell in row:

                if isinstance(cell, str) and cell.startswith(tuple(prefixes)):

                    matching_values.append(cell)

        return matching_values

 

    # Main function to copy values matching specific prefixes from an Excel file to clipboard

    def copy_matching_values_to_clipboard(file_path, prefixes):

        # Check if the file exists

        if not os.path.exists(file_path):

            print(f"Error: The file '{file_path}' does not exist.")

            return

 

        try:

            workbook = openpyxl.load_workbook(file_path)

            all_matching_values = []

 

            # Iterate through all sheets

            for sheet_name in workbook.sheetnames:

                sheet = workbook[sheet_name]

                matching_values = filter_values(sheet, prefixes)

                all_matching_values.extend(matching_values)

 

            # Convert list of matching values into a DataFrame format

            df_values = pd.DataFrame(all_matching_values, columns=["0"])

 

            # Copy the DataFrame to the clipboard (Excel-style)

            df_values.to_clipboard(index=False, header=True)

            print(f"Values matching {prefixes} copied to clipboard successfully!")

 

             # Print clipboard contents

            clipboard_contents = pyperclip.paste()  # Get contents of the clipboard

 

            #print("Clipboard contents:")

            #print(clipboard_contents)

 

        except Exception as e:

            print(f"An error occurred: {e}")

 

    # File path

    file_path = r"C:\Users\svc.cost.user\Desktop\W16 Budget Overview - 2025.xlsx"

 

    # Prefixes to filter

    prefixes = ["DEV2", "DV25", "BD25"]

 

    # Run the function

    copy_matching_values_to_clipboard(file_path, prefixes)

 

    # Proceed with the SAP interactions for exporting data

    SAP_OBJ.session.findById("wnd[0]/usr/txt$6-KOKRS").caretPosition = (4)

    SAP_OBJ.session.findById("wnd[0]/usr/btn%__6ORDGRP_%_APP_%-VALU_PUSH").press()

    SAP_OBJ.session.findById("wnd[1]/tbar[0]/btn[24]").press()

    SAP_OBJ.session.findById("wnd[1]/tbar[0]/btn[8]").press()

    SAP_OBJ.session.findById("wnd[0]/tbar[1]/btn[8]").press()

 

    # Collapse fields, double click total value and click 'Orders: Actual Line Items' line

    SAP_OBJ.session.findById("wnd[0]/mbar/menu[3]/menu[1]").select()

    SAP_OBJ.session.findById("wnd[0]/usr/lbl[5,8]").setFocus()

    SAP_OBJ.session.findById("wnd[0]/usr/lbl[5,8]").caretPosition = (9)

    SAP_OBJ.session.findById("wnd[0]").sendVKey(2)

    SAP_OBJ.session.findById("wnd[1]/usr/lbl[1,2]").setFocus()

    SAP_OBJ.session.findById("wnd[1]/usr/lbl[1,2]").caretPosition = (22)

    SAP_OBJ.session.findById("wnd[1]").sendVKey(2)

 

    # Filter to get Inventory usage UIIL types

    SAP_OBJ.session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell").setCurrentCell(-1, "CEL_KTXT")

    SAP_OBJ.session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell").selectColumn("CEL_KTXT")

    SAP_OBJ.session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell").contextMenu()

    SAP_OBJ.session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell").selectContextMenuItem("&FILTER")

    SAP_OBJ.session.findById("wnd[1]/usr/ssub%_SUBSCREEN_FREESEL:SAPLSSEL:1105/ctxt%%DYN001-LOW").text = "Inventory Usage UIIL"

    SAP_OBJ.session.findById("wnd[1]/usr/ssub%_SUBSCREEN_FREESEL:SAPLSSEL:1105/ctxt%%DYN001-LOW").caretPosition = (20)

    SAP_OBJ.session.findById("wnd[1]/tbar[0]/btn[0]").press()

 

    SAP_OBJ.session.findById("wnd[0]/mbar/menu[0]/menu[3]/menu[1]").select()

 

    #Define the directory and file name

    SAP_OBJ.session.findById("wnd[1]/usr/ctxtDY_PATH").text = Dir

    SAP_OBJ.session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = ("WT2_Recharges_2025.XLSX")

    SAP_OBJ.session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = (12)

    #Replace file, save

    SAP_OBJ.session.findById("wnd[1]/tbar[0]/btn[11]").press()

    #Exit transaction, Go back to main screen

    SAP_OBJ.session.findById("wnd[0]/tbar[0]/btn[12]").press()

    SAP_OBJ.session.findById("wnd[0]/tbar[0]/btn[12]").press()

    SAP_OBJ.session.findById("wnd[1]/usr/btnBUTTON_YES").press()

    SAP_OBJ.session.findById("wnd[0]/tbar[0]/btn[12]").press()

    print(f'WT2 Recharges Complete')









 #Production Order Extract Block

    #Open COOIS

    SAP_OBJ.session.FindById("wnd[0]/tbar[0]/okcd").Text = "/ncoois"

    SAP_OBJ.session.FindById("wnd[0]").SendVKey (0)

    #Enter Layout - this layout will work with the code

    SAP_OBJ.session.FindById("wnd[0]/usr/ssub%_SUBSCREEN_TOPBLOCK:PPIO_ENTRY:1100/ctxtPPIO_ENTRY_SC1100-ALV_VARIANT").Text = "/SPOTFIRE1"

    #Enter Start and End Date Into SAP (Actual release date)

    SAP_OBJ.session.FindById("wnd[0]/usr/tabsTABSTRIP_SELBLOCK/tabpSEL_00/ssub%_SUBSCREEN_SELBLOCK:PPIO_ENTRY:1200/ctxtS_ISTFR-LOW").Text = ProdOSD

    SAP_OBJ.session.FindById("wnd[0]/usr/tabsTABSTRIP_SELBLOCK/tabpSEL_00/ssub%_SUBSCREEN_SELBLOCK:PPIO_ENTRY:1200/ctxtS_ISTFR-HIGH").Text = ProdOEd

    SAP_OBJ.session.FindById("wnd[0]/usr/tabsTABSTRIP_SELBLOCK/tabpSEL_00/ssub%_SUBSCREEN_SELBLOCK:PPIO_ENTRY:1200/ctxtS_ISTST-HIGH").SetFocus()

    #SAP_OBJ.session.FindById("wnd[0]/usr/tabsTABSTRIP_SELBLOCK/tabpSEL_00/ssub%_SUBSCREEN_SELBLOCK:PPIO_ENTRY:1200/ctxtS_ISTST-HIGH").caretPosition = 10

    #Execute transaction

    SAP_OBJ.session.FindById("wnd[0]/tbar[1]/btn[8]").Press()

    SAP_OBJ.session.FindById("wnd[1]/usr/btnBUTTON_2").Press()

    #Export to excel

    SAP_OBJ.session.FindById("wnd[0]/usr/cntlCUSTOM/shellcont/shell/shellcont/shell").PressToolbarContextButton ("&MB_EXPORT")

    SAP_OBJ.session.FindById("wnd[0]/usr/cntlCUSTOM/shellcont/shell/shellcont/shell").SelectContextMenuItem ("&XXL")

    #SAP_OBJ.session.FindById("wnd[1]/tbar[0]/btn[0]").Press()

    #Set the file name and file path to save the file

    SAP_OBJ.session.FindById("wnd[1]/usr/ctxtDY_PATH").Text = Dir

    SAP_OBJ.session.FindById("wnd[1]/usr/ctxtDY_FILENAME").Text = ProdOExpName

    SAP_OBJ.session.FindById("wnd[1]/tbar[0]/btn[11]").Press()

    print('Production/COOIS Complete')






# Operations Start Here

    # Import required modules/classes

    import pandas as pd

 

    # Erase data from the Export Process Sheet

    df = pd.read_excel(r"\\groups-server\groups\Design\Dept Only\Cost Analysts\PlanVsActual DATA\NowCast Data Auto\Export Process.xlsx")

    df.iloc[0:, :] = ""

    df.to_excel(r"\\groups-server\groups\Design\Dept Only\Cost Analysts\PlanVsActual DATA\NowCast Data Auto\Export Process.xlsx", index=False)

    # Copy and Paste data from ProdOrder Legacy Document and combine with SAP export

    df1 = pd.read_excel(r"G:\Engineering\Public\EPM\Cost Visualisation Dashboard\Build_Codes\Inventory Recoveries\ProductionOrders_2020ToDate.xlsx")

    df2 = pd.read_excel(r"\\groups-server\groups\Design\Dept Only\Cost Analysts\PlanVsActual DATA\NowCast Data Auto\ProdOrder-Latest_AUTO.xlsx")

    data_to_copy = df1.iloc[1:, :14]

    df2 = pd.concat([df2, data_to_copy], ignore_index=True)

    # Remove specific order types from the DataFrame

    df2 = df2[~df2['Order Type'].isin(['Z005', 'Z006', 'Z009', 'ZBDV'])]

 

    # Convert the 'Release date (actual)' column to datetime type

    df2['Release date (actual)'] = pd.to_datetime(df2['Release date (actual)'])

    # Filter the DataFrame based on the 'Release date (actual)' greater than or equal to OpFiltDate

    df2 = df2.loc[df2['Release date (actual)'] >= OpFiltDate]

 

    # Save the updated DataFrame to the Export Process Sheet

    df2.to_excel(r"\\groups-server\groups\Design\Dept Only\Cost Analysts\PlanVsActual DATA\NowCast Data Auto\Export Process.xlsx", index=False)

 

    # Copy List Of Orders to Clipboard

 

    # Read the data from the Export Process Sheet

    df = pd.read_excel(r"\\groups-server\groups\Design\Dept Only\Cost Analysts\PlanVsActual DATA\NowCast Data Auto\Export Process.xlsx")

 

    # Select the data from the first column and drop any rows with NaN (missing) values

    data_to_copy = df.iloc[:, 0]

    data_to_copy.dropna(inplace=True)

 

    # Copy the selected data to the clipboard (to be used elsewhere in the script)

    data_to_copy.to_clipboard(index=False, header=False)

   

    # SAP interaction for Operations

    # Set the SAP transaction code and execute it

    SAP_OBJ.session.FindById("wnd[0]/tbar[0]/okcd").Text = "/ncoois"

    SAP_OBJ.session.FindById("wnd[0]").SendVKey(0)

 

    # Select the relevant list type and ALV variant for Operations

    SAP_OBJ.session.FindById("wnd[0]/usr/ssub%_SUBSCREEN_TOPBLOCK:PPIO_ENTRY:1100/cmbPPIO_ENTRY_SC1100-PPIO_LISTTYP").Key = "PPIOO000" #Set Order headers to operations

    SAP_OBJ.session.FindById("wnd[0]/usr/ssub%_SUBSCREEN_TOPBLOCK:PPIO_ENTRY:1100/ctxtPPIO_ENTRY_SC1100-ALV_VARIANT").Text = "/PZ" #Set the layout to PZ as this works with the code

 

    # Set the focus and caret position for the ALV variant field

    SAP_OBJ.session.FindById("wnd[0]/usr/ssub%_SUBSCREEN_TOPBLOCK:PPIO_ENTRY:1100/ctxtPPIO_ENTRY_SC1100-ALV_VARIANT").SetFocus()

    SAP_OBJ.session.FindById("wnd[0]/usr/ssub%_SUBSCREEN_TOPBLOCK:PPIO_ENTRY:1100/ctxtPPIO_ENTRY_SC1100-ALV_VARIANT").caretPosition = (3)

    # Press buttons to navigate to the required data and export it to Excel

    SAP_OBJ.session.FindById("wnd[0]/usr/tabsTABSTRIP_SELBLOCK/tabpSEL_00/ssub%_SUBSCREEN_SELBLOCK:PPIO_ENTRY:1200/btn%_S_AUFNR_%_APP_%-VALU_PUSH").Press()

    SAP_OBJ.session.FindById("wnd[1]/tbar[0]/btn[24]").Press()

    SAP_OBJ.session.FindById("wnd[1]/tbar[0]/btn[8]").Press()

    SAP_OBJ.session.FindById("wnd[0]/tbar[1]/btn[8]").Press()

    SAP_OBJ.session.FindById("wnd[1]/usr/btnBUTTON_2").Press()

    # Export the data to Excel using the "XXL" format

    SAP_OBJ.session.FindById("wnd[0]/usr/cntlCUSTOM/shellcont/shell/shellcont/shell").PressToolbarButton("&NAVIGATION_PROFILE_TOOLBAR_EXPAND")

    SAP_OBJ.session.FindById("wnd[0]/usr/cntlCUSTOM/shellcont/shell/shellcont/shell").PressToolbarContextButton("&MB_EXPORT")

    SAP_OBJ.session.FindById("wnd[0]/usr/cntlCUSTOM/shellcont/shell/shellcont/shell").SelectContextMenuItem("&XXL")

    #SAP_OBJ.session.FindById("wnd[1]/tbar[0]/btn[0]").Press()

    SAP_OBJ.session.FindById("wnd[1]/usr/ctxtDY_PATH").Text = Dir

    SAP_OBJ.session.FindById("wnd[1]/usr/ctxtDY_FILENAME").Text = OpExpName

    SAP_OBJ.session.FindById("wnd[1]/tbar[0]/btn[11]").Press()

    print('Operations Complete')






# Components Starts Here

    # Import data from the "Export Process Sheet" Excel file and erase its contents

    import pandas as pd

 

    df = pd.read_excel(r"\\groups-server\groups\Design\Dept Only\Cost Analysts\PlanVsActual DATA\NowCast Data Auto\Export Process.xlsx")

    df.iloc[0:, :] = ""  # Erase data in the DataFrame

    df.to_excel(r"\\groups-server\groups\Design\Dept Only\Cost Analysts\PlanVsActual DATA\NowCast Data Auto\Export Process.xlsx", index=False)

   

    # Copy and Paste data from "ProdOrder Legacy Document" and combine it with the SAP export

    df1 = pd.read_excel(r"G:\Engineering\Public\EPM\Cost Visualisation Dashboard\Build_Codes\Inventory Recoveries\ProductionOrders_2020ToDate.xlsx")

    df2 = pd.read_excel(r"\\groups-server\groups\Design\Dept Only\Cost Analysts\PlanVsActual DATA\NowCast Data Auto\ProdOrder-Latest_AUTO.xlsx")

    data_to_copy = df1.iloc[1:, :14]

    df2 = pd.concat([df2, data_to_copy], ignore_index=True)

    df2 = df2[~df2['Order Type'].isin(['Z005', 'Z006', 'Z009', 'ZBDV'])]

    df2['Release date (actual)'] = pd.to_datetime(df2['Release date (actual)'])

    df2 = df2.loc[df2['Release date (actual)'] >= ComFiltDate]

    df2.to_excel(r"\\groups-server\groups\Design\Dept Only\Cost Analysts\PlanVsActual DATA\NowCast Data Auto\Export Process.xlsx", index=False)

   

    # Copy List Of Orders to Clipboard

    df = pd.read_excel(r"\\groups-server\groups\Design\Dept Only\Cost Analysts\PlanVsActual DATA\NowCast Data Auto\Export Process.xlsx")

    data_to_copy = df.iloc[:, 0]

    data_to_copy.dropna(inplace=True)

    data_to_copy.to_clipboard(index=False, header=False)

   

    # SAP Interaction for Components

 

    # Set the SAP transaction code and execute it

    SAP_OBJ.session.FindById("wnd[0]/tbar[0]/okcd").Text = "/ncoois"

    SAP_OBJ.session.FindById("wnd[0]").SendVKey(0)

 

    # Select the relevant list type and ALV variant for Components - List: Components, Layout /PZ

    SAP_OBJ.session.FindById("wnd[0]/usr/ssub%_SUBSCREEN_TOPBLOCK:PPIO_ENTRY:1100/cmbPPIO_ENTRY_SC1100-PPIO_LISTTYP").Key = "PPIOM000"

    SAP_OBJ.session.FindById("wnd[0]/usr/ssub%_SUBSCREEN_TOPBLOCK:PPIO_ENTRY:1100/ctxtPPIO_ENTRY_SC1100-ALV_VARIANT").Text = "/PZ"

 

    # Set the focus and caret position for the ALV variant field

    SAP_OBJ.session.FindById("wnd[0]/usr/ssub%_SUBSCREEN_TOPBLOCK:PPIO_ENTRY:1100/ctxtPPIO_ENTRY_SC1100-ALV_VARIANT").SetFocus()

    SAP_OBJ.session.FindById("wnd[0]/usr/ssub%_SUBSCREEN_TOPBLOCK:PPIO_ENTRY:1100/ctxtPPIO_ENTRY_SC1100-ALV_VARIANT").caretPosition = (3)

 

    # Press buttons to navigate to the required data and export it to Excel

    SAP_OBJ.session.FindById("wnd[0]/usr/tabsTABSTRIP_SELBLOCK/tabpSEL_00/ssub%_SUBSCREEN_SELBLOCK:PPIO_ENTRY:1200/btn%_S_AUFNR_%_APP_%-VALU_PUSH").Press()

    SAP_OBJ.session.FindById("wnd[1]/tbar[0]/btn[24]").Press()

    SAP_OBJ.session.FindById("wnd[1]/tbar[0]/btn[8]").Press()

    SAP_OBJ.session.FindById("wnd[0]/tbar[1]/btn[8]").Press()

    SAP_OBJ.session.FindById("wnd[1]/usr/btnBUTTON_2").Press()

 

    # Expand the toolbar and initiate the export in "XXL" format

    SAP_OBJ.session.FindById("wnd[0]/usr/cntlCUSTOM/shellcont/shell/shellcont/shell").PressToolbarButton("&NAVIGATION_PROFILE_TOOLBAR_EXPAND")

    SAP_OBJ.session.FindById("wnd[0]/usr/cntlCUSTOM/shellcont/shell/shellcont/shell").PressToolbarContextButton("&MB_EXPORT")

    SAP_OBJ.session.FindById("wnd[0]/usr/cntlCUSTOM/shellcont/shell/shellcont/shell").SelectContextMenuItem("&XXL")

 

    # Confirm the export

    #SAP_OBJ.session.FindById("wnd[1]/tbar[0]/btn[0]").Press()

 

    # Set the path and filename for the exported data in Excel

    SAP_OBJ.session.FindById("wnd[1]/usr/ctxtDY_PATH").Text = Dir

    SAP_OBJ.session.FindById("wnd[1]/usr/ctxtDY_FILENAME").Text = ComExpName

 

    # Press the "Save" button to complete the Excel export

    SAP_OBJ.session.FindById("wnd[1]/tbar[0]/btn[11]").Press()

    print('Components Complete')




    #Start Of ExOP Extraction

    # Start Of ExOP Extraction

 

    # Import pandas library for data manipulation

    import pandas as pd

 

    # Read the data from the ExtOp.xlsx file

    df = pd.read_excel(r"\\groups-server\groups\Design\Dept Only\Cost Analysts\PlanVsActual DATA\NowCast Data Auto\ExtOp.xlsx")

 

    # Erase the data in the DataFrame by setting all values to an empty string

    df.iloc[0:, :] = ""

 

    # Save the empty DataFrame back to the ExtOp.xlsx file

    df.to_excel(r"\\groups-server\groups\Design\Dept Only\Cost Analysts\PlanVsActual DATA\NowCast Data Auto\ExtOp.xlsx", index=False)

 

    # Read data from multiple files and merge them into a single DataFrame

    df1 = pd.read_excel(r"\\groups-server\groups\Design\Dept Only\Cost Analysts\PlanVsActual DATA\NowCast Data Auto\Operations Latest_Auto.xlsx")

    df2 = pd.read_excel(r"\\groups-server\groups\Design\Dept Only\Cost Analysts\PlanVsActual DATA\NowCast Data Auto\Legacy\Operations 20.xlsx")

    df3 = pd.read_excel(r"\\groups-server\groups\Design\Dept Only\Cost Analysts\PlanVsActual DATA\NowCast Data Auto\Legacy\Operations 21.xlsx")

    df4 = pd.read_excel(r"\\groups-server\groups\Design\Dept Only\Cost Analysts\PlanVsActual DATA\NowCast Data Auto\Legacy\Operations 22.xlsx")

    df5 = pd.read_excel(r"\\groups-server\groups\Design\Dept Only\Cost Analysts\PlanVsActual DATA\NowCast Data Auto\Legacy\Operations 23.xlsx")

    df6 = pd.read_excel(r"\\groups-server\groups\Design\Dept Only\Cost Analysts\PlanVsActual DATA\NowCast Data Auto\Legacy\Operations 24 01.01 - 30.06.xlsx")

 

    # Concatenate the DataFrames vertically to merge them into one

    df = pd.concat([df1, df2, df3, df4, df5,df6], axis=0)

 

    # Remove rows where the 'Purchasing Document' column is not empty

    df = df[df['Purchasing Document'].notna()]

 

    # Save the modified DataFrame back to the ExtOp.xlsx file

    df.to_excel(r"\\groups-server\groups\Design\Dept Only\Cost Analysts\PlanVsActual DATA\NowCast Data Auto\ExtOp.xlsx", index=False)

 

    # Perform SAP interactions for ExtOP extraction

    SAP_OBJ.session.FindById("wnd[0]/tbar[0]/okcd").Text = "/nzpoe"

    SAP_OBJ.session.FindById("wnd[0]").SendVKey(0)

 

    # Set the selection criteria for the SAP data extraction, PO created between the 'Latest' date field.

    SAP_OBJ.session.FindById("wnd[0]/usr/ctxtS_EINDT-LOW").Text = ExOpSD

    SAP_OBJ.session.FindById("wnd[0]/usr/ctxtS_EINDT-HIGH").Text = ExOpEd

    SAP_OBJ.session.FindById("wnd[0]/usr/ctxtS_BSART-LOW").Text = "NB"

    SAP_OBJ.session.FindById("wnd[0]/usr/ctxtS_BSART-LOW").SetFocus()

    SAP_OBJ.session.FindById("wnd[0]/usr/ctxtS_BSART-LOW").caretPosition = (2)

 

    # Read the data from the ExtOp.xlsx file

    df = pd.read_excel(r"\\groups-server\groups\Design\Dept Only\Cost Analysts\PlanVsActual DATA\NowCast Data Auto\ExtOp.xlsx")

 

    # Extract the 'Order' data and copy it to clipboard

    order_data = df['Order'].drop_duplicates().to_clipboard(index=False, header=False)

 

    # Perform SAP interactions for further ExOP extraction

    SAP_OBJ.session.FindById("wnd[0]/usr/btn%_S_AUFNR_%_APP_%-VALU_PUSH").Press()

    SAP_OBJ.session.FindById("wnd[1]/tbar[0]/btn[24]").Press()

    SAP_OBJ.session.FindById("wnd[1]/tbar[0]/btn[8]").Press()

 

    # Read the data from the ExtOp.xlsx file again

    df = pd.read_excel(r"\\groups-server\groups\Design\Dept Only\Cost Analysts\PlanVsActual DATA\NowCast Data Auto\ExtOp.xlsx")

 

    # Extract the 'Purchasing Document' data and copy it to clipboard

    order_data = df['Purchasing Document'].drop_duplicates().to_clipboard(index=False, header=False)

 

    # SAP interactions for ExOP extraction

 

    # Perform SAP interaction to press the button for selecting the value of a field

    SAP_OBJ.session.FindById("wnd[0]/usr/btn%_S_EBELN_%_APP_%-VALU_PUSH").Press()

 

    # Perform SAP interaction to press a button in the toolbar of a pop-up window

    SAP_OBJ.session.FindById("wnd[1]/tbar[0]/btn[24]").Press()

 

    # Perform SAP interaction to press a button in the toolbar of a pop-up window

    SAP_OBJ.session.FindById("wnd[1]/tbar[0]/btn[8]").Press()

 

    # Untick boxes to make ZPOE work

    SAP_OBJ.session.FindById("wnd[0]/usr/chkCB_IRFQ").Selected = (False)

    SAP_OBJ.session.FindById("wnd[0]/usr/chkCB_IDCF").Selected = (False)

    SAP_OBJ.session.FindById("wnd[0]/usr/chkCB_IRO1").Selected = (False)

    SAP_OBJ.session.FindById("wnd[0]/usr/chkCB_ICOR").Selected = (False)

 

    # Perform SAP interaction to set the text value of layout field to "/PZ"

    SAP_OBJ.session.FindById("wnd[0]/usr/ctxtP_VARI").Text = "/PZ"

    SAP_OBJ.session.FindById("wnd[0]/usr/ctxtP_VARI").SetFocus()

    # Perform SAP interaction to set the caret position in a field (cursor position)

    SAP_OBJ.session.FindById("wnd[0]/usr/ctxtP_VARI").caretPosition = (3)

    SAP_OBJ.session.FindById("wnd[0]/tbar[1]/btn[8]").Press()

    SAP_OBJ.session.FindById("wnd[0]/tbar[1]/btn[46]").Press()

    SAP_OBJ.session.FindById("wnd[0]/tbar[1]/btn[43]").Press()

    # Select Excel file

    #SAP_OBJ.session.FindById("wnd[1]/tbar[0]/btn[0]").Press()

    # Perform SAP interaction to set the text value of a field to 'Dir' variable

    SAP_OBJ.session.FindById("wnd[1]/usr/ctxtDY_PATH").Text = Dir

    # Perform SAP interaction to set the text value of a field to 'ExOPExpName' variable

    SAP_OBJ.session.FindById("wnd[1]/usr/ctxtDY_FILENAME").Text = ExOPExpName

    # Save The Excel FIle

    SAP_OBJ.session.FindById("wnd[1]/tbar[0]/btn[11]").Press()

    print('ExtOP Complete')





    #Settlement ZPSD Export

    #Read file, filter to remove certain order types, then copy order column to clipboard

    df1 = pd.read_excel(r"\\groups-server\groups\Design\Dept Only\Cost Analysts\PlanVsActual DATA\NowCast Data Auto\ProdOrder-Latest_AUTO.xlsx")

    df1 = df1[~df1['Order Type'].isin(['Z005','Z006','Z009','ZBDV'])]

   

    df2 = pd.read_excel(r"\\groups-server\groups\Design\Dept Only\Cost Analysts\PlanVsActual DATA\NowCast Data Auto\Legacy\Prod Ord 2024 01.01-30.06.xlsx")

    df2 = df2[~df2['Order Type'].isin(['Z005','Z006','Z009','ZBDV'])]

    df1 = pd.concat([df1,df2], ignore_index=True)

 

    #Count number of rows. ZPSD transaction too big in SAP, split in two files to export and then merge with legacy.

    row_count = len(df1)

    #First ZPSD Latest Export

    data_to_copy = df1.iloc[:(row_count//2), 0]

    data_to_copy.dropna(inplace=True)  # Remove rows with NaN values

    data_to_copy.to_clipboard(index=False, header=False)

    #ZPSD Transaction

    #Perform SAP interactions to extract settlement data

    SAP_OBJ.session.FindById("wnd[0]/tbar[0]/okcd").Text = "/nzpsd"

    SAP_OBJ.session.FindById("wnd[0]").SendVKey (0)

    SAP_OBJ.session.FindById("wnd[0]/usr/btn%_S_AUFNR_%_APP_%-VALU_PUSH").Press()

    SAP_OBJ.session.FindById("wnd[1]/tbar[0]/btn[24]").Press()

    SAP_OBJ.session.FindById("wnd[1]/tbar[0]/btn[8]").Press()

    SAP_OBJ.session.FindById("wnd[0]/tbar[1]/btn[8]").Press()

    SAP_OBJ.session.FindById("wnd[0]/usr/cntlCCONT/shellcont/shell/shellcont/shell").PressToolbarContextButton ("&MB_EXPORT")

    SAP_OBJ.session.FindById("wnd[0]/usr/cntlCCONT/shellcont/shell/shellcont/shell").SelectContextMenuItem ("&XXL")

    #REMOVE#SAP_OBJ.session.FindById("wnd[1]/tbar[0]/btn[0]").Press()

    SAP_OBJ.session.FindById("wnd[1]/usr/ctxtDY_PATH").Text = Dir

    SAP_OBJ.session.FindById("wnd[1]/usr/ctxtDY_FILENAME").Text = "Settlement Latest_Auto1.xlsx "

    SAP_OBJ.session.FindById("wnd[1]/tbar[0]/btn[11]").Press()

 

    #Second ZPSD Latest Export

    data_to_copy = df1.iloc[(row_count//2):row_count, 0]

    data_to_copy.dropna(inplace=True)  # Remove rows with NaN values

    data_to_copy.to_clipboard(index=False, header=False)

    #ZPSD Transaction

    #Perform SAP interactions to extract settlement data

    SAP_OBJ.session.FindById("wnd[0]/tbar[0]/okcd").Text = "/nzpsd"

    SAP_OBJ.session.FindById("wnd[0]").SendVKey (0)

    SAP_OBJ.session.FindById("wnd[0]/usr/btn%_S_AUFNR_%_APP_%-VALU_PUSH").Press()

    SAP_OBJ.session.FindById("wnd[1]/tbar[0]/btn[24]").Press()

    SAP_OBJ.session.FindById("wnd[1]/tbar[0]/btn[8]").Press()

    SAP_OBJ.session.FindById("wnd[0]/tbar[1]/btn[8]").Press()

    SAP_OBJ.session.FindById("wnd[0]/usr/cntlCCONT/shellcont/shell/shellcont/shell").PressToolbarContextButton ("&MB_EXPORT")

    SAP_OBJ.session.FindById("wnd[0]/usr/cntlCCONT/shellcont/shell/shellcont/shell").SelectContextMenuItem ("&XXL")

    #REMOVE#SAP_OBJ.session.FindById("wnd[1]/tbar[0]/btn[0]").Press()

    SAP_OBJ.session.FindById("wnd[1]/usr/ctxtDY_PATH").Text = Dir

    SAP_OBJ.session.FindById("wnd[1]/usr/ctxtDY_FILENAME").Text = "Settlement Latest_Auto2.xlsx "

    SAP_OBJ.session.FindById("wnd[1]/tbar[0]/btn[11]").Press()

   

    # Read excel files

    df1 = pd.read_excel(r"\\groups-server\groups\Design\Dept Only\Cost Analysts\PlanVsActual DATA\NowCast Data Auto\Settlement Latest_Auto1.xlsx")

    df2 = pd.read_excel(r"\\groups-server\groups\Design\Dept Only\Cost Analysts\PlanVsActual DATA\NowCast Data Auto\Settlement Latest_Auto2.xlsx")

    df3 = pd.read_excel(r"\\groups-server\groups\Design\Dept Only\Cost Analysts\PlanVsActual DATA\NowCast Data Auto\Legacy\Settlement Legacy All.xlsx")

   

    #Concatenate together, save to Settlements 2020todate_auto

    df = pd.concat([df1, df2,df3], ignore_index=True)

    df.to_excel(r"\\groups-server\groups\Design\Dept Only\Cost Analysts\PlanVsActual DATA\NowCast Data Auto\Settle 2020todate_Auto.xlsx", index=False)

    print('Settlement ZPSD Export Complete')    

   

   

 

#ZMUD8 - 2025

    # Start Of ZMUD8 Extract 2025

    # Set SAP transaction code to access ZMUD8

    SAP_OBJ.session.FindById("wnd[0]/tbar[0]/okcd").Text = "/nzmud8"

    # Press Enter to execute the transaction

    SAP_OBJ.session.FindById("wnd[0]").SendVKey (0)

    # Set the start date for the ZMUD8 extraction

    SAP_OBJ.session.FindById("wnd[0]/usr/ctxtS_ERDAT-LOW").Text = ZMudSD

    # Set the end date for the ZMUD8 extraction

    SAP_OBJ.session.FindById("wnd[0]/usr/ctxtS_ERDAT-HIGH").Text = ZMudEd

    # Move the focus to the end date field

    SAP_OBJ.session.FindById("wnd[0]/usr/ctxtS_ERDAT-HIGH").SetFocus()

    # Set the caret position to 10 (to select the day part of the date)

    SAP_OBJ.session.FindById("wnd[0]/usr/ctxtS_ERDAT-HIGH").caretPosition = 10

    # Press the 'Execute' button to start the extraction

    SAP_OBJ.session.FindById("wnd[0]/tbar[1]/btn[8]").Press()

    # Set the current cell to the 'ERDAT' column in the ALV Grid

    SAP_OBJ.session.FindById("wnd[0]/usr/cntlCCONT/shellcont/shell/shellcont/shell").SetCurrentCell, r"-1", ("ERDAT")

    # Select the 'ERDAT' column in the ALV Grid

    SAP_OBJ.session.FindById("wnd[0]/usr/cntlCCONT/shellcont/shell/shellcont/shell").SelectColumn ("ERDAT")

    # Open the context menu in the ALV Grid

    SAP_OBJ.session.FindById("wnd[0]/usr/cntlCCONT/shellcont/shell/shellcont/shell").ContextMenu()

    # Sort the data in ascending order based on the 'ERDAT' column

    SAP_OBJ.session.FindById("wnd[0]/usr/cntlCCONT/shellcont/shell/shellcont/shell").PressToolbarContextButton ("&SORT_ASC")

    # Press the 'Export' button in the ALV Grid

    SAP_OBJ.session.FindById("wnd[0]/usr/cntlCCONT/shellcont/shell/shellcont/shell").PressToolbarContextButton ("&MB_EXPORT")

    # Select the 'XXL' option for export

    SAP_OBJ.session.FindById("wnd[0]/usr/cntlCCONT/shellcont/shell/shellcont/shell").SelectContextMenuItem ("&XXL")

    # Press the 'Export to Spreadsheet' button to start the export

    #SAP_OBJ.session.FindById("wnd[1]/tbar[0]/btn[0]").Press()

    # Set the path for the exported file

    SAP_OBJ.session.FindById("wnd[1]/usr/ctxtDY_PATH").Text = Dir

    # Set the filename for the exported file

    SAP_OBJ.session.FindById("wnd[1]/usr/ctxtDY_FILENAME").Text = ZMudExpName

    # Press the 'Save' button to save the exported file

    SAP_OBJ.session.FindById("wnd[1]/tbar[0]/btn[11]").Press()






#Start Of ZMB51 Extract

 

    SAP_OBJ.session.findById("wnd[0]/tbar[0]/okcd").text = "/nzmb51"

    SAP_OBJ.session.findById("wnd[0]").sendVKey (0)

    SAP_OBJ.session.findById("wnd[0]/usr/ctxtS_WERKS-LOW").text = "1000"

    SAP_OBJ.session.findById("wnd[0]/usr/ctxtS_MATNR-LOW").text = ""

    SAP_OBJ.session.findById("wnd[0]/usr/ctxtS_MATNR-HIGH").text = ""

    SAP_OBJ.session.findById("wnd[0]/usr/ctxtS_BWART-LOW").text = ""

    SAP_OBJ.session.findById("wnd[0]/usr/ctxtS_BWART-HIGH").text = ""

    SAP_OBJ.session.findById("wnd[0]/usr/ctxtS_LGORT-LOW").text = ""

    SAP_OBJ.session.findById("wnd[0]/usr/ctxtS_LGORT-HIGH").text = ""

    SAP_OBJ.session.findById("wnd[0]/usr/ctxtS_UMLGO-LOW").text = ""

    SAP_OBJ.session.findById("wnd[0]/usr/ctxtS_UMLGO-HIGH").text = ""

    SAP_OBJ.session.findById("wnd[0]/usr/ctxtS_BUDAT-LOW").text = MatMSD

    SAP_OBJ.session.findById("wnd[0]/usr/ctxtS_BUDAT-HIGH").text = MatMEd

    SAP_OBJ.session.findById("wnd[0]/usr/radRB_ALV").setFocus()

    SAP_OBJ.session.findById("wnd[0]/tbar[1]/btn[8]").press()

    SAP_OBJ.session.findById("wnd[0]/usr/cntlCCONT/shellcont/shell/shellcont/shell").pressToolbarContextButton ("&MB_EXPORT")

    SAP_OBJ.session.findById("wnd[0]/usr/cntlCCONT/shellcont/shell/shellcont/shell").selectContextMenuItem ("&PC")

    SAP_OBJ.session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]").select()

    SAP_OBJ.session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]").setFocus()

    SAP_OBJ.session.findById("wnd[1]/tbar[0]/btn[0]").press()

    SAP_OBJ.session.FindById("wnd[1]/usr/ctxtDY_PATH").Text = Dir

    SAP_OBJ.session.FindById("wnd[1]/usr/ctxtDY_FILENAME").Text = MatExpName

    SAP_OBJ.session.FindById("wnd[1]/tbar[0]/btn[11]").Press()

    print('ZMB51 Extract Complete')







#Start of ZLLR Extract

 

    # Read data from the ZMUD8-Latest_AUTO.xlsx file

    df = pd.read_excel(r"\\groups-server\groups\Design\Dept Only\Cost Analysts\PlanVsActual DATA\NowCast Data Auto\ZMUD8-Latest_AUTO.xlsx")

 

    # Filter data where 'Matl Usage Doc Year' contains '2025'

    df = df[df['Matl Usage Doc Year'].astype(str).str.contains('2025')]

 

    # Keep only the 'Component' column and drop other columns

    df.drop(df.columns.difference(['Component']), axis=1, inplace=True)

 

    # Drop duplicate rows based on the 'Component' column, keeping the first occurrence

    df.drop_duplicates(subset='Component', keep='first', inplace=True)

 

    # Copy list of 'Component' values to clipboard

    ZMUD = df['Component'].drop_duplicates().to_clipboard(index=False, header=False)

 

    # Set SAP transaction code to access ZLLR

    SAP_OBJ.session.FindById("wnd[0]/tbar[0]/okcd").Text = "/nzllr"

    # Press Enter to execute the transaction

    SAP_OBJ.session.FindById("wnd[0]").SendVKey (0)

 

    # Press buttons in the SAP GUI

    SAP_OBJ.session.FindById("wnd[0]/usr/btn%_S_MATNR_%_APP_%-VALU_PUSH").Press()

    SAP_OBJ.session.FindById("wnd[1]/tbar[0]/btn[24]").Press()

    SAP_OBJ.session.FindById("wnd[1]/tbar[0]/btn[8]").Press()

    SAP_OBJ.session.FindById("wnd[0]/tbar[1]/btn[8]").Press()

    SAP_OBJ.session.FindById("wnd[0]/mbar/menu[0]/menu[1]/menu[2]").Select()

    SAP_OBJ.session.FindById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]").Select()

    SAP_OBJ.session.FindById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]").SetFocus()

    SAP_OBJ.session.FindById("wnd[1]/tbar[0]/btn[0]").Press()

 

    # Set the path and filename for the exported file (using the ZLLRExpName variable)

    SAP_OBJ.session.FindById("wnd[1]/usr/ctxtDY_PATH").Text = Dir

    SAP_OBJ.session.FindById("wnd[1]/usr/ctxtDY_FILENAME").Text = ZLLRExpName

    # Set focus on the path field and position the cursor at the end of the path

    SAP_OBJ.session.FindById("wnd[1]/usr/ctxtDY_PATH").SetFocus()

    SAP_OBJ.session.FindById("wnd[1]/usr/ctxtDY_PATH").caretPosition = (42)

    # Press the 'Save' button to save the exported file

    SAP_OBJ.session.FindById("wnd[1]/tbar[0]/btn[11]").Press()

 

    # Log out of SAP

    SAP_OBJ.session.findById("wnd[0]/mbar/menu[3]/menu[12]").select()

    SAP_OBJ.session.findById("wnd[1]/usr/btnSPOP-OPTION1").press()

 

    # Calculate elapsed time in minutes

    elapsed_time = time.time() - start_time

    elapsed_time = elapsed_time / 60

    elapsed_time = round(elapsed_time, 2)

 

    # Create the body of the email with the extraction status and elapsed time

    body = "Extraction Successful. Elapsed Time: " + str(elapsed_time) + " Minutes."

 

    # Attempt to send an email with the extraction status

    send_email("[INFO] NowCast Data Extraction Successful", body, "adad@mercedesamgf1.com;rvudathu@mercedesamgf1.com")

 

except Exception as e:

    # If there is an exception, send an email indicating that the extraction failed

    send_email("[INFO] Nowcast Data Extraction Failed", "NowCast Extraction has failed, please contact Alisha Dad", "adad@mercedesamgf1.com;rvudathu@mercedesamgf1.com")

 

# Import necessary modules

import win32com.client

import pythoncom

import psutil

 

# Function to quit Excel by killing the process

def quit_excel():

    for process in psutil.process_iter():

        if process.name() == "EXCEL.EXE":

            process.kill()

 

# Call the function to quit Excel

quit_excel()
